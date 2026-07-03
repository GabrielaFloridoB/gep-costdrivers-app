import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="GEP – Cost Drivers Database Export",
    page_icon="📊",
    layout="wide",
)

# ── Styling helpers ────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
GLOSSARY_FILL = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
DATA_FONT    = Font(name="Arial", size=9)
ALT_FILL     = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
THIN_BORDER  = Border(
    left=Side(style="thin", color="B8CCE4"),
    right=Side(style="thin", color="B8CCE4"),
    top=Side(style="thin", color="B8CCE4"),
    bottom=Side(style="thin", color="B8CCE4"),
)

COLUMN_WIDTHS = {
    "ID": 10,
    "Indicator": 70,
    "Source": 25,
    "Country": 20,
    "HS Code": 18,
    "ImportExport": 20,
    "Update frequency": 22,
    "Group": 22,
    "LastDateUpdated": 20,
    "Last forecast update": 22,
    "Frequency of publication source": 35,
}

COLUMN_GLOSSARY = [
    ("ID",                              "Unique code that identifies each record in the table. Used to unambiguously reference individual rows."),
    ("Indicator",                       "Name of the monitored indicator — e.g. commodity price, import volume, ocean freight rate."),
    ("Source",                          "Data origin: the entity or platform responsible for publishing the indicator (e.g. UN Comtrade, World Bank, Eurostat)."),
    ("Country",                         "Country the indicator refers to — may be the country of origin, destination, or reference for the measurement."),
    ("HS Code",                         "Harmonized System code: classifies the product for international trade purposes. A globally adopted standard for customs and trade."),
    ("ImportExport",                    "Indicates whether the indicator's trade flow is import or export, defining the direction of the commercial flow being analyzed."),
    ("Update frequency platform",       "How often does the internal platform update the data for this indicator (e.g., daily, weekly, monthly)? This depends on the availability of data from the source."),
    ("Group",                           "Thematic category of the indicator (e.g. energy, metals, grains, freight). Helps with filtering and browsing by topic."),
    ("LastDateUpdated",                 "Date of the last actual data update on the platform. Indicates how fresh the available information is."),
    ("Last forecast update",            "Date of the last update to the forecast or projection linked to the indicator. Relevant when the field includes forward-looking data."),
    ("Frequency of publication source", "The frequency with which the original source publishes its data (e.g., monthly, quarterly, annually) may differ from the frequency with which the platform itself updates its data, depending on the publication of new data."),
]

# ── Excel builder ──────────────────────────────────────────────────────────────
def build_excel(df: pd.DataFrame) -> bytes:
    wb = openpyxl.Workbook()

    # ── Sheet 1: result ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "result"
    ws.freeze_panes = "A2"

    headers = list(df.columns)
    ws.append(headers)

    # Header row style
    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(col_name, 18)

    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = DATA_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # Auto-filter on headers
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── Sheet 2: Column glossary ───────────────────────────────────────────────
    wg = wb.create_sheet("Column glossary")
    wg.freeze_panes = "A2"

    wg.column_dimensions["A"].width = 35
    wg.column_dimensions["B"].width = 120

    wg.append(["Column", "Description"])
    for col_idx in range(1, 3):
        cell = wg.cell(row=1, column=col_idx)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill      = GLOSSARY_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER

    for r_idx, (col_name, desc) in enumerate(COLUMN_GLOSSARY, 2):
        c1 = wg.cell(row=r_idx, column=1, value=col_name)
        c2 = wg.cell(row=r_idx, column=2, value=desc)
        for c in (c1, c2):
            c.font      = DATA_FONT
            c.border    = THIN_BORDER
            c.alignment = Alignment(vertical="center", wrap_text=True)
        if r_idx % 2 == 0:
            c1.fill = ALT_FILL
            c2.fill = ALT_FILL

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── DB connection helper ───────────────────────────────────────────────────────
@st.cache_data(show_spinner=False, ttl=300)
def load_from_db(
    host: str, port: int, database: str,
    user: str, password: str,
    table: str, filters: dict
) -> pd.DataFrame:
    """Connect to MySQL-compatible DB and fetch costdrivers_infos."""
    import sqlalchemy

    url = sqlalchemy.engine.URL.create(
        drivername="mysql+pymysql",
        username=user,
        password=password,
        host=host,
        port=port,
        database=database,
    )
    engine = sqlalchemy.create_engine(url)

    where_clauses = []
    params = {}

    if filters.get("groups"):
        placeholders = ", ".join(f":g{i}" for i in range(len(filters["groups"])))
        where_clauses.append(f"`Group` IN ({placeholders})")
        for i, g in enumerate(filters["groups"]):
            params[f"g{i}"] = g

    if filters.get("countries"):
        placeholders = ", ".join(f":c{i}" for i in range(len(filters["countries"])))
        where_clauses.append(f"Country IN ({placeholders})")
        for i, c in enumerate(filters["countries"]):
            params[f"c{i}"] = c

    if filters.get("sources"):
        placeholders = ", ".join(f":s{i}" for i in range(len(filters["sources"])))
        where_clauses.append(f"Source IN ({placeholders})")
        for i, s in enumerate(filters["sources"]):
            params[f"s{i}"] = s

    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""
    query = f"SELECT * FROM {table} {where_sql}"

    with engine.connect() as conn:
        df = pd.read_sql(sqlalchemy.text(query), conn, params=params)

    return df


# ── UI ─────────────────────────────────────────────────────────────────────────
st.title("📊 GEP – Cost Drivers Database Export")
st.caption("Conecte ao banco, filtre os dados e gere o Excel padronizado.")

# ── Sidebar: connection settings ───────────────────────────────────────────────
with st.sidebar:
    st.header("🔌 Conexão com o Banco")

    host     = st.text_input("Host",     value="localhost")
    port     = st.number_input("Porta",  value=3306, min_value=1, max_value=65535)
    database = st.text_input("Database", value="costdrivers_dg_prod")
    user     = st.text_input("Usuário",  value="")
    password = st.text_input("Senha",    type="password")
    table    = st.text_input(
        "Tabela",
        value="`00_raw`.costdrivers_infos",
        help="Ex.: `00_raw`.costdrivers_infos"
    )

    st.divider()
    st.header("🔍 Filtros")

    group_filter   = st.text_area("Groups (um por linha)", placeholder="Services\nFuels\nMinerals")
    country_filter = st.text_area("Countries (um por linha)", placeholder="Brazil\nUnited States")
    source_filter  = st.text_area("Sources (um por linha)", placeholder="US BLS\nFRED/US BLS")

    load_btn = st.button("🔄 Carregar dados do banco", use_container_width=True)

# ── Main area ──────────────────────────────────────────────────────────────────
tab_db, tab_upload = st.tabs(["🗄️ Buscar do Banco", "📁 Usar arquivo local"])

# ── Tab 1: from DB ─────────────────────────────────────────────────────────────
with tab_db:
    if load_btn:
        if not user or not password:
            st.warning("Preencha usuário e senha na barra lateral.")
        else:
            filters = {
                "groups":    [g.strip() for g in group_filter.splitlines() if g.strip()],
                "countries": [c.strip() for c in country_filter.splitlines() if c.strip()],
                "sources":   [s.strip() for s in source_filter.splitlines() if s.strip()],
            }
            with st.spinner("Conectando ao banco e carregando dados…"):
                try:
                    df = load_from_db(host, int(port), database, user, password, table, filters)
                    st.session_state["db_df"] = df
                    st.success(f"✅ {len(df):,} registros carregados.")
                except Exception as e:
                    st.error(f"Erro ao conectar: {e}")

    if "db_df" in st.session_state:
        df = st.session_state["db_df"]

        # Preview & metrics
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Total de registros", f"{len(df):,}")
        col2.metric("Países",    df["Country"].nunique() if "Country" in df.columns else "–")
        col3.metric("Groups",    df["Group"].nunique()   if "Group"   in df.columns else "–")
        col4.metric("Fontes",    df["Source"].nunique()  if "Source"  in df.columns else "–")

        with st.expander("👀 Pré-visualização (primeiras 100 linhas)"):
            st.dataframe(df.head(100), use_container_width=True)

        # Generate Excel
        st.divider()
        if st.button("📥 Gerar Excel padronizado", use_container_width=True):
            with st.spinner("Gerando Excel…"):
                xlsx_bytes = build_excel(df)
                date_str = datetime.today().strftime("%d-%m-%Y")
                filename = f"GEP_-_Data_Base_{date_str}.xlsx"

            st.download_button(
                label=f"⬇️ Baixar {filename}",
                data=xlsx_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    else:
        st.info("Configure a conexão na barra lateral e clique em **Carregar dados do banco**.")

# ── Tab 2: local file ──────────────────────────────────────────────────────────
with tab_upload:
    st.markdown("Faça upload de um Excel ou CSV existente para regenerá-lo no formato padrão GEP.")
    uploaded = st.file_uploader("Selecione o arquivo", type=["xlsx", "xls", "csv"])

    if uploaded:
        with st.spinner("Lendo arquivo…"):
            if uploaded.name.endswith(".csv"):
                df_up = pd.read_csv(uploaded)
            else:
                df_up = pd.read_excel(uploaded, sheet_name="result")

        st.success(f"✅ {len(df_up):,} registros lidos.")

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Total de registros", f"{len(df_up):,}")
        col2.metric("Países",  df_up["Country"].nunique() if "Country" in df_up.columns else "–")
        col3.metric("Groups",  df_up["Group"].nunique()   if "Group"   in df_up.columns else "–")
        col4.metric("Fontes",  df_up["Source"].nunique()  if "Source"  in df_up.columns else "–")

        with st.expander("👀 Pré-visualização"):
            st.dataframe(df_up.head(100), use_container_width=True)

        if st.button("📥 Gerar Excel padronizado (arquivo local)", use_container_width=True):
            with st.spinner("Gerando Excel…"):
                xlsx_bytes = build_excel(df_up)
                date_str = datetime.today().strftime("%d-%m-%Y")
                filename = f"GEP_-_Data_Base_{date_str}.xlsx"

            st.download_button(
                label=f"⬇️ Baixar {filename}",
                data=xlsx_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
